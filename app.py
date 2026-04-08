import os
import re
import shutil
import unicodedata
import zipfile
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import streamlit as st


# ============================================================
#   DEFINICIÓN DEL TEMPLATE MODELO (reconstruido en código)
# ============================================================
SHEET_NAME = "ImportacionesEvweb"

FORMATO_NUMERO = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'
FORMATO_FECHA  = 'm/d/yy h:mm'

# (nombre, ancho, num_format, h_align, v_align, header_num_format, header_h_align, header_v_align)
COLUMNAS = [
    ("matricula",              7.89,     "General",     None,     None,    "General",         None,     None),
    ("afiliado_numero",        15.0,     "General",     "center", None,    "0",               None,     None),
    ("afiliado_denominacion",  41.66,    "General",     "left",   None,    "General",         "left",   None),
    ("fecha_prestacion",       18.55,    FORMATO_FECHA, "center", "center","@",               "center", None),
    ("cod_practica",           11.66,    "General",     None,     None,    "General",         None,     None),
    ("cantidad",               None,     "General",     None,     None,    "General",         None,     None),
    ("actuacion",              None,     "General",     None,     None,    "General",         None,     None),
    ("tipo_facturacion",       11.55,    "General",     None,     None,    "General",         None,     None),
    ("honorario",              13.22,    FORMATO_NUMERO,None,     None,    "General",         None,     None),
    ("1er_ayudante",           11.55,    "General",     None,     None,    "General",         None,     None),
    ("2do_ayudante",           None,     "General",     None,     None,    "General",         None,     None),
    ("gasto",                  11.66,    "General",     None,     None,    "General",         None,     None),
    ("modulo",                 11.55,    "General",     None,     None,    "General",         None,     None),
    ("aparatoligia",           None,     "General",     None,     None,    "General",         None,     None),
    ("id_anticipo",            None,     "General",     None,     None,    "General",         None,     None),
    ("iva",                    11.66,    "General",     "center", None,    "General",         None,     None),
    ("numaut",                 15.33,    "General",     "center", None,    "General",         None,     None),
    ("cuit",                   13.66,    "General",     None,     None,    "General",         None,     None),
    ("fecha_presentacion",     16.78,    FORMATO_FECHA, "center", "center",FORMATO_FECHA,     "center", "center"),
    ("coseguro",               11.66,    FORMATO_NUMERO,None,     None,    "General",         None,     None),
    ("modalidadCoseguro",      11.55,    "General",     None,     None,    "General",         None,     None),
    ("nroAutorizacion",        11.66,    "General",     None,     None,    "General",         None,     None),
]

NOMBRES_COLUMNAS = [c[0] for c in COLUMNAS]

FUENTE_BASE              = Font(name="MS Sans Serif", size=10, bold=False)
FUENTE_HEADER            = Font(name="MS Sans Serif", size=10, bold=True)
FUENTE_HEADER_FECHA_PRES = Font(name="MS Sans Serif", size=10, bold=False)


def crear_template_vacio():
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for idx, meta in enumerate(COLUMNAS, start=1):
        (nombre, ancho, num_fmt, h_align, v_align,
         header_num_fmt, header_h_align, header_v_align) = meta

        letra = get_column_letter(idx)
        if ancho is not None:
            ws.column_dimensions[letra].width = ancho

        celda = ws.cell(row=1, column=idx, value=nombre)
        celda.font = FUENTE_HEADER_FECHA_PRES if nombre == "fecha_presentacion" else FUENTE_HEADER
        celda.number_format = header_num_fmt
        if header_h_align or header_v_align:
            celda.alignment = Alignment(horizontal=header_h_align, vertical=header_v_align)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}1"
    return wb


def aplicar_estilo_fila_datos(ws, fila_idx):
    for col_idx, meta in enumerate(COLUMNAS, start=1):
        (nombre, ancho, num_fmt, h_align, v_align, *_rest) = meta
        celda = ws.cell(row=fila_idx, column=col_idx)
        celda.font = FUENTE_BASE
        celda.number_format = num_fmt
        if h_align or v_align:
            celda.alignment = Alignment(horizontal=h_align, vertical=v_align)


# ============================================================
#                MAPEO DE COLUMNAS DE ENTRADA
# ============================================================
MAPEO = {
    "matricula":             ["cuenta", "cuenta_matricula", "matricula", "n° de cuenta", "nro de cuenta", "numero de cuenta"],
    "afiliado_numero":       ["credencial", "numero de afiliado", "nro afiliado", "afiliado", "afiliado n°", "codigo_afiliado", "credencial socio"],
    "afiliado_denominacion": ["apellido_afiliado", "nombre de afiliado", "afiliado", "nom.beneficiario", "apellido y nombre socio", "apellido y nombre"],
    "fecha_prestacion":      ["fecha_transaccion", "fecha prestacion", "fecha consulta", "fecha de transaccion", "fecha turno", "f. trans.", "fecha", "TURNO","fecha transaccion"],
    "cod_practica":          ["prestación", "prestacion", "codigo", "código", "practica", "práctica", "cod prest", "cod_prest"],
    "cantidad":              ["cantidad", "cant", "cant."],
    "honorario":             ["total", "importe total", "importe_total"],
    "iva":                   ["iva_template", "iva t", "iva p", "iva"],
    "numaut":                ["transaccion_item", "numero autorizacion", "nro.trans.", "nro trans", "id", "id transaccion", "id transaciion", "NRO. ORDEN"],
    "coseguro":              ["copago", "coseguro"],
}

COLUMNAS_CERO   = ["actuacion", "1er_ayudante", "2do_ayudante", "gasto", "modulo", "aparatoligia"]
COLUMNAS_VACIAS = ["tipo_facturacion", "id_anticipo", "cuit", "modalidadCoseguro", "nroAutorizacion"]

FILAS_POR_TEMPLATE = 1500


def _quitar_acentos(texto):
    """Elimina tildes y diacríticos."""
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(ch for ch in nfkd if not unicodedata.combining(ch))


def _normalizar(texto):
    if texto is None:
        return ""
    t = str(texto).strip().lower()
    t = _quitar_acentos(t)
    return re.sub(r"[\s_\-\.]+", "", t)


def _buscar_columna(df_cols_norm, candidatos):
    for cand in candidatos:
        cand_n = _normalizar(cand)
        for real, norm in df_cols_norm.items():
            if norm == cand_n:
                return real
    for cand in candidatos:
        cand_n = _normalizar(cand)
        for real, norm in df_cols_norm.items():
            if cand_n and cand_n in norm:
                return real
    return None


def _ultimo_dia_mes_anterior(hoy=None):
    hoy = hoy or datetime.now()
    primer_dia = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return (primer_dia - pd.Timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)


def _extraer_numero(valor):
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    digitos = re.sub(r"\D", "", str(valor))
    if not digitos:
        return None
    try:
        return int(digitos)
    except ValueError:
        return None


def construir_filas_salida(df):
    cols_norm = {c: _normalizar(c) for c in df.columns}
    resolver = {dest: _buscar_columna(cols_norm, cands) for dest, cands in MAPEO.items()}

    # Log de mapeo para verificación
    st.write("\n🔎 Mapeo de columnas detectado:")
    for dest, origen in resolver.items():
        estado = f"← '{origen}'" if origen else "← (no encontrada)"
        st.write(f"   {dest:25s} {estado}")
    st.write("")

    fecha_pres = _ultimo_dia_mes_anterior()
    filas = []

    for _, row in df.iterrows():
        fila = {}

        col = resolver["matricula"]
        fila["matricula"] = row[col] if col else None

        col = resolver["afiliado_numero"]
        fila["afiliado_numero"] = _extraer_numero(row[col]) if col else None

        col = resolver["afiliado_denominacion"]
        fila["afiliado_denominacion"] = row[col] if col else None

        col = resolver["fecha_prestacion"]
        val = row[col] if col else None
        if val is not None and not (isinstance(val, float) and pd.isna(val)):
            try:
                fila["fecha_prestacion"] = pd.to_datetime(val).to_pydatetime()
            except Exception:
                fila["fecha_prestacion"] = val
        else:
            fila["fecha_prestacion"] = None

        col = resolver["cod_practica"]
        fila["cod_practica"] = row[col] if col else None

        col = resolver["cantidad"]
        fila["cantidad"] = row[col] if col else None

        for c in COLUMNAS_CERO:
            fila[c] = 0
        for c in COLUMNAS_VACIAS:
            fila[c] = None

        col = resolver["honorario"]
        fila["honorario"] = row[col] if col else None

        col = resolver["iva"]
        fila["iva"] = row[col] if col else 0

        col = resolver["numaut"]
        fila["numaut"] = _extraer_numero(row[col]) if col else None

        fila["fecha_presentacion"] = fecha_pres

        col = resolver["coseguro"]
        fila["coseguro"] = row[col] if col else 0

        filas.append(fila)
    return filas


def escribir_template(filas, ruta_salida):
    wb = crear_template_vacio()
    ws = wb.active

    for i, fila in enumerate(filas, start=2):
        for col_idx, nombre in enumerate(NOMBRES_COLUMNAS, start=1):
            ws.cell(row=i, column=col_idx, value=fila.get(nombre))
        aplicar_estilo_fila_datos(ws, i)

    wb.save(ruta_salida)
    wb.close()


def generar_zip_templates(uploaded_file_obj, ruta_zip_salida,
                          filas_por_template=FILAS_POR_TEMPLATE):
    # Use BytesIO to read the uploaded file directly without saving to disk first
    df = pd.read_excel(uploaded_file_obj, dtype=object)
    filas = construir_filas_salida(df)

    # Use BytesIO for templates to avoid disk writes if possible, or temporary directory
    # For this, we'll stick to a temporary directory as openpyxl's save needs a path
    carpeta_tmp = "_tmp_templates"
    os.makedirs(carpeta_tmp, exist_ok=True)
    rutas = []

    total = len(filas)
    if total == 0:
        ruta = os.path.join(carpeta_tmp, "Template_01.xlsx")
        escribir_template([], ruta)
        rutas.append(ruta)
    else:
        n = (total + filas_por_template - 1) // filas_por_template
        for idx in range(n):
            ini = idx * filas_por_template
            fin = ini + filas_por_template
            nombre = f"Template_{idx + 1:02d}.xlsx"
            ruta = os.path.join(carpeta_tmp, nombre)
            escribir_template(filas[ini:fin], ruta)
            rutas.append(ruta)

    with zipfile.ZipFile(ruta_zip_salida, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in rutas:
            zf.write(r, arcname=os.path.basename(r))

    # Clean up temporary files/directory
    shutil.rmtree(carpeta_tmp, ignore_errors=True)
    
    return ruta_zip_salida, total, len(rutas)


# ============================================================
#                  STREAMLIT APP LAYOUT
# ============================================================
st.set_page_config(layout="wide")
st.title("Generador de Templates Excel")
st.markdown("Sube tu archivo Excel y genera templates automáticamente.")

uploaded_file = st.file_uploader("📤 Sube tu archivo Excel (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    st.write(f"📄 Archivo recibido: {uploaded_file.name}")

    try:
        ruta_zip, total_filas, n_templates = generar_zip_templates(
            uploaded_file_obj=uploaded_file,
            ruta_zip_salida="Templates.zip",
        )

        st.success(f"✅ {total_filas} filas procesadas en {n_templates} template(s)")
        
        with open(ruta_zip, "rb") as f:
            st.download_button(
                label="⬇️ Descargar Templates.zip",
                data=f.read(),
                file_name="Templates.zip",
                mime="application/zip"
            )
        os.remove(ruta_zip) # Clean up the generated zip file

    except Exception as e:
        st.error(f"Ocurrió un error al procesar el archivo: {e}")
        st.exception(e)

